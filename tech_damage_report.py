#!/usr/bin/env python3
"""
TechReplacement — Automated OTDR Damage Report
===============================================
Reads raw OTDR traces from the EXFO proprietary block (RawSamples) the same
way a tech would look at the screen, then produces a damage report document
matching the format your boss expects.

For each fiber the script:
  1. Reads the RawSamples trace from ExfoNewProprietaryBlock (correct distances)
  2. Self-calibrates the distance axis using the launch connector Fresnel peak
  3. Fits the Rayleigh backscatter baseline (robust linear fit)
  4. Detects the noise floor — where the fiber goes blind
  5. Finds breaks: strong Fresnel spikes above baseline, confirmed by short EOL
  6. Precisely locates each break (Fresnel peak vs. splice table position)
  7. Measures segment attenuation slopes to detect bend-loss zones
  8. For fibers with breaks, switches to B-direction data past the break (blind fill)
  9. Writes an Excel report matching splice_report.xlsx format

Output format matches splice_report.xlsx:
  - Row 1 : splice distances (km)
  - Row 2 : column headers (dark navy background, white text)
  - Row 3+ : one row per 12-fiber ribbon
  - Pink cell   : normal splice loss entry  — "180 .162"
  - Red cell    : break with offset         — "180 BREAK 3.624 (50ft from splice)"
  - Orange cell : confirmed break/blind     — "180 broke"
  - Blue cell   : B-direction fill          — "180 .162 (B)"
  - Yellow col  : bend-loss zones           — "37-48 +0.06 dB/km"

Usage:
    python3 tech_damage_report.py <dir_a> <dir_b>
    python3 tech_damage_report.py <dir_a> <dir_b> --site-a EUG --site-b SAL
    python3 tech_damage_report.py <dir_a> <dir_b> --threshold 0.15 --output damage.xlsx
    python3 tech_damage_report.py <dir_a> <dir_b> --ribbon-size 12 --bend-threshold 0.05
"""

import argparse
import os
import sys
import zlib
import struct
import numpy as np
from collections import defaultdict, Counter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from sor_reader324802a import (
    _parse_block_directory, _read_ior, _parse_key_events
)

# ─── Constants ────────────────────────────────────────────────────────────────
RIBBON_SIZE        = 12
MIN_POP_SPLICE     = 20
SPLICE_MATCH_KM    = 1.5
NOISE_STD_THRESH   = 0.08    # dB rolling std to declare noise floor
NOISE_WINDOW       = 60      # samples for rolling std
RIBBON_EOL_TOL_KM  = 0.5     # fibers shorter than ribbon median by this → broken
STEP_LOSS_THRESH   = 1.0     # dB drop in trace to call a step-loss at break site
C_LIGHT            = 2.998e8

# Fresnel reflection height above extrapolated Rayleigh baseline at a break.
# A clean glass-air interface (fresh dig/cut) creates a strong Fresnel spike.
# A crushed fiber has no clean surface — the trace steps down with no spike.
REFL_CLEAN_BREAK   = 5.0    # dB above baseline → clean break (fresh dig / aerial cut)
REFL_PARTIAL_BREAK = 2.0    # dB above baseline → partial break / stress fracture
REFL_CRUSH_MIN     = 0.5    # dB above baseline → weak trace of Fresnel → crush w/ gap


# ─────────────────────────────────────────────────────────────────────────────
#  Raw trace reading
# ─────────────────────────────────────────────────────────────────────────────

def read_rawsamples(filepath):
    """
    Read EXFO RawSamples from ExfoNewProprietaryBlock.
    Returns (raw_db, dist_km, dx_km, pre_trigger_km, IOR) or None.
    raw_db: float64, EXFO convention — higher = stronger signal, slopes down.
    """
    with open(filepath, 'rb') as f:
        data = f.read()

    IOR    = _read_ior(data)
    blocks = _parse_block_directory(data)

    blk_name = next((n for n in blocks if 'ExfoNewProprietaryBlock' in n), None)
    if blk_name is None:
        return None

    exfo  = blocks[blk_name]
    chunk = data[exfo['body']:exfo['offset'] + exfo['size']]

    s0 = None
    for off in (40, 36):
        try:
            s0 = zlib.decompress(chunk[off:])
            break
        except zlib.error:
            pass
    if s0 is None:
        return None

    rs_off = s0.find(b'RawSamples')
    if rs_off < 0:
        return None

    avail = len(s0) - (rs_off + 15)
    n     = avail // 2
    if n < 100:
        return None

    raw    = np.frombuffer(s0[rs_off + 15:rs_off + 15 + n * 2],
                           dtype='<u2').astype(np.float64) / 1000.0
    dx_km  = 50e-9 * C_LIGHT / (2 * IOR) / 1000.0
    i_conn = int(np.argmax(raw[:400]))
    pre_km = i_conn * dx_km
    dist   = np.arange(n, dtype=np.float64) * dx_km - pre_km

    return raw, dist, dx_km, pre_km, IOR


# ─────────────────────────────────────────────────────────────────────────────
#  Noise floor detection
# ─────────────────────────────────────────────────────────────────────────────

def find_noise_floor_km(dist, raw_db):
    """
    Return the km position where the trace enters the noise floor.
    Uses rolling standard deviation: when std > NOISE_STD_THRESH the signal
    is lost.  Starts searching after the first 20% of the trace.
    """
    n       = len(raw_db)
    start_i = max(int(n * 0.15), NOISE_WINDOW)

    for i in range(start_i, n - NOISE_WINDOW):
        window_std = np.std(raw_db[i:i + NOISE_WINDOW])
        if window_std > NOISE_STD_THRESH:
            return float(dist[i])

    return float(dist[-1])


# ─────────────────────────────────────────────────────────────────────────────
#  Rayleigh baseline fitting
# ─────────────────────────────────────────────────────────────────────────────

def fit_baseline(dist, raw_db, exclude_kms, pulse_km=0.12, noise_floor_km=None):
    """
    Fit a linear Rayleigh baseline, excluding windows around known events
    and the noise-floor region.  Returns a numpy poly1d callable.
    """
    end_km = (noise_floor_km * 0.95) if noise_floor_km else (dist[-1] * 0.95)

    mask = (dist > 0.8) & (dist < end_km)
    for ek in exclude_kms:
        mask &= np.abs(dist - ek) > pulse_km * 1.5

    if mask.sum() < 20:
        mask = (dist > 0.8) & (dist < end_km)

    x, y = dist[mask], raw_db[mask]

    # First pass
    coeffs = np.polyfit(x, y, 1)
    resid  = y - np.polyval(coeffs, x)
    # Reject outliers > 2σ (spike / event remnants)
    keep   = np.abs(resid) < 2.0 * np.std(resid)
    if keep.sum() > 15:
        coeffs = np.polyfit(x[keep], y[keep], 1)

    return np.poly1d(coeffs)


# ─────────────────────────────────────────────────────────────────────────────
#  Break detection — ribbon-based EOL comparison
# ─────────────────────────────────────────────────────────────────────────────

def collect_eof_positions(fibers):
    """Return {fnum: eof_km} for all fibers that have an end-of-fiber event."""
    eof = {}
    for fnum, r in fibers.items():
        for e in r['events']:
            if e['is_end']:
                eof[fnum] = e['dist_km']
                break
    return eof


def detect_breaks_by_ribbon(eof_a, ribbon_size=RIBBON_SIZE, tol_km=RIBBON_EOL_TOL_KM):
    """
    Primary break detection: compare each fiber's EOL to its ribbon-mates.

    Fibers in the same ribbon are wound together at the same helix angle so
    they have nearly identical optical path lengths.  A fiber that ends
    significantly shorter than its ribbon-mates has a break, regardless of
    where on the span that break is.

    Uses the top 75% of EOL values in each ribbon as the "healthy" reference
    so that broken fibers don't pull the median down.

    Returns {fnum: rough_break_km} for all detected broken fibers.
    """
    # Group EOLs by ribbon
    ribbon_eofs = defaultdict(list)
    for fnum, eof_km in eof_a.items():
        ribbon_idx = (fnum - 1) // ribbon_size
        ribbon_eofs[ribbon_idx].append((fnum, eof_km))

    broken = {}
    for ribbon_idx, fiber_eofs in ribbon_eofs.items():
        if len(fiber_eofs) < 2:
            continue

        # Sort descending; use top 75% as healthy reference
        sorted_eofs = sorted([v for _, v in fiber_eofs], reverse=True)
        n_ref       = max(1, int(len(sorted_eofs) * 0.75))
        ribbon_ref  = float(np.median(sorted_eofs[:n_ref]))

        for fnum, eof_km in fiber_eofs:
            if eof_km < ribbon_ref - tol_km:
                broken[fnum] = eof_km   # break is at/near the EOL position

    return broken


# ─────────────────────────────────────────────────────────────────────────────
#  Break localization — step-loss in raw trace
# ─────────────────────────────────────────────────────────────────────────────

def localize_break_step(raw_db, dist, dx_km, pre_km, rough_km):
    """
    Localize a break by finding the sharp downward step in the raw trace
    near rough_km (the EOL position from the event table).

    A real break shows as the trace dropping below the extrapolated Rayleigh
    line — NOT necessarily a spike upward.  We fit the pre-break slope and
    find the first sample where the trace drops more than STEP_LOSS_THRESH
    below that extrapolation.

    Returns (precise_break_km, step_loss_db) or (rough_km, None).
    """
    # Convert rough_km to sample index
    i_eof     = int(round((rough_km + pre_km) / dx_km))
    fit_back  = int(2.0 / dx_km)    # 2 km of pre-break trace for slope fit
    srch_back = int(1.0 / dx_km)    # search window: 1 km before EOL
    srch_fwd  = int(0.3 / dx_km)    # and 0.3 km after

    fit_start = max(5, i_eof - fit_back)
    fit_end   = max(fit_start + 10, i_eof - srch_back)
    srch_s    = max(fit_end, i_eof - srch_back)
    srch_e    = min(len(raw_db) - 1, i_eof + srch_fwd)

    if fit_end - fit_start < 10 or srch_s >= srch_e:
        return rough_km, None

    x_fit    = np.arange(fit_start, fit_end, dtype=float)
    y_fit    = raw_db[fit_start:fit_end]
    coeffs   = np.polyfit(x_fit, y_fit, 1)

    x_srch   = np.arange(srch_s, srch_e, dtype=float)
    y_extrap = np.polyval(coeffs, x_srch)
    y_actual = raw_db[srch_s:srch_e]
    drop     = y_extrap - y_actual   # positive = trace is below expected

    # First sample where the drop exceeds threshold = break position
    for i, d in enumerate(drop):
        if d >= STEP_LOSS_THRESH:
            step_idx   = srch_s + i
            precise_km = float(dist[step_idx])
            return round(precise_km, 5), round(float(d), 2)

    # No clear step found — return rough position
    return rough_km, None


# ─────────────────────────────────────────────────────────────────────────────
#  Direct trace loss measurement — replaces event-table splice_loss lookup
# ─────────────────────────────────────────────────────────────────────────────

def measure_event_loss_from_trace(raw_db, dist, dx_km, pre_km, event_km,
                                   noise_km=None,
                                   other_event_kms=None,
                                   pre_window_km=1.2,
                                   post_window_km=1.2,
                                   guard_km=0.18):
    """
    Measure splice/event loss directly from the raw backscatter trace.

    This replicates what a tech does visually: fit Rayleigh lines through the
    trace on each side of the event and read off the step height.  Using the
    raw trace instead of the OTDR's auto-calculated splice_loss value gives:
      - Loss values at positions the OTDR didn't record as events
      - Better accuracy when nearby events contaminate the OTDR's window
      - Independent validation of the OTDR's own numbers

    The fit windows avoid a ±guard_km dead-zone around the event itself
    and shrink away from other nearby events so they don't contaminate the
    slope estimate.

    Returns (loss_db, pre_slope_db_km, post_slope_db_km) or (None, None, None).
    loss_db is positive for a downward step (real loss), negative for gain
    (which means the OTDR is seeing an artifact or the event is noise).
    """
    i_event = int(round((event_km + pre_km) / dx_km))
    i_guard = max(1, int(guard_km / dx_km))
    i_win   = max(5, int(pre_window_km / dx_km))
    i_noise = int((noise_km + pre_km) / dx_km) if noise_km else len(raw_db) - 1

    # Build pre-event window: (event - guard) going back pre_window_km
    pre_end   = i_event - i_guard
    pre_start = max(5, pre_end - i_win)

    # Build post-event window: (event + guard) going forward post_window_km
    post_start = i_event + i_guard
    post_end   = min(i_noise - 5, post_start + i_win)

    # Trim windows away from other known events
    if other_event_kms:
        for ek in other_event_kms:
            i_ek = int(round((ek + pre_km) / dx_km))
            if i_ek < i_event - i_guard:   # event is before → clip pre window
                pre_start = max(pre_start, i_ek + i_guard)
            elif i_ek > i_event + i_guard: # event is after  → clip post window
                post_end = min(post_end, i_ek - i_guard)

    if pre_end - pre_start < 8 or post_end - post_start < 8:
        return None, None, None
    if post_start >= i_noise - 5:
        return None, None, None

    def robust_fit(x_arr, y_arr):
        coeffs = np.polyfit(x_arr, y_arr, 1)
        resid  = y_arr - np.polyval(coeffs, x_arr)
        keep   = np.abs(resid) < 2.0 * np.std(resid)
        if keep.sum() >= 6:
            coeffs = np.polyfit(x_arr[keep], y_arr[keep], 1)
        return coeffs

    x_pre = np.arange(pre_start, pre_end, dtype=float)
    x_post = np.arange(post_start, post_end, dtype=float)

    coeffs_pre  = robust_fit(x_pre,  raw_db[pre_start:pre_end])
    coeffs_post = robust_fit(x_post, raw_db[post_start:post_end])

    # Extrapolate both lines to the event sample
    val_pre  = np.polyval(coeffs_pre,  float(i_event))
    val_post = np.polyval(coeffs_post, float(i_event))

    loss_db          = val_pre - val_post            # positive = downward step = loss
    # Convert slope from dB/sample to dB/km
    pre_slope_db_km  = abs(float(coeffs_pre[0]))  / dx_km
    post_slope_db_km = abs(float(coeffs_post[0])) / dx_km

    return round(loss_db, 4), round(pre_slope_db_km, 4), round(post_slope_db_km, 4)


def scan_for_missed_events(raw_db, dist, dx_km, pre_km, noise_km,
                            known_event_kms, min_step_db=0.12):
    """
    Scan the raw trace for step losses the OTDR event table missed.

    Uses a robust two-window approach: fit Rayleigh slope on 0.8 km each
    side of a candidate position, extrapolate both to the midpoint, and
    measure the step.  Requires the step to exceed min_step_db AND the
    post-window slope to be continuous (not just a noise spike).

    Returns list of {km, step_db} for candidate missed events.
    Note: caller should apply ribbon-consensus filtering to suppress noise.
    """
    noise_margin_km = max(0.0, noise_km - 1.0)
    i_noise  = int((noise_margin_km + pre_km) / dx_km)
    i_start  = max(5, int((2.0 + pre_km) / dx_km))
    half_win = max(5, int(0.8 / dx_km))
    step_win = max(2, int(0.15 / dx_km))   # narrow comparison window at the step

    missed = []

    # Only check at coarse resolution (every 0.1 km) to reduce false positives
    step = max(1, int(0.1 / dx_km))
    i = i_start + half_win
    while i < i_noise - half_win:
        # Narrow window straddling the candidate position
        pre_end  = i - step_win
        pre_start= max(5, i - half_win)
        post_start = i + step_win
        post_end = min(i_noise - 1, i + half_win)

        if pre_end - pre_start < 5 or post_end - post_start < 5:
            i += step
            continue

        x_pre  = np.arange(pre_start, pre_end, dtype=float)
        x_post = np.arange(post_start, post_end, dtype=float)
        try:
            c_pre  = np.polyfit(x_pre,  raw_db[pre_start:pre_end], 1)
            c_post = np.polyfit(x_post, raw_db[post_start:post_end], 1)
        except Exception:
            i += step
            continue

        val_pre  = np.polyval(c_pre,  float(i))
        val_post = np.polyval(c_post, float(i))
        drop = val_pre - val_post

        if drop >= min_step_db:
            km = float(dist[i])
            if not any(abs(km - ek) < SPLICE_MATCH_KM for ek in known_event_kms):
                missed.append({'km': round(km, 3), 'step_db': round(drop, 3)})
                i += half_win    # skip past event
                continue
        i += step

    return missed


def apply_ribbon_consensus_to_missed(all_missed_raw, ribbon_size=RIBBON_SIZE,
                                      min_fibers=2, cluster_km=0.5):
    """
    Filter raw per-fiber missed events through ribbon consensus.

    An event that appears in only one fiber at a given km is almost certainly
    trace noise.  An event appearing in ≥ min_fibers fibers of the same ribbon
    at roughly the same distance is real.

    Returns {fnum: [{km, step_db}]} containing only consensus-validated events.
    """
    # Build: ribbon_idx -> list of (fnum, km, step_db)
    ribbon_candidates = defaultdict(list)
    for fnum, evs in all_missed_raw.items():
        ri = (fnum - 1) // ribbon_size
        for ev in evs:
            ribbon_candidates[ri].append((fnum, ev['km'], ev['step_db']))

    validated = defaultdict(list)
    for ri, candidates in ribbon_candidates.items():
        if len(candidates) < min_fibers:
            continue
        # Cluster candidates that are within cluster_km of each other
        candidates.sort(key=lambda x: x[1])   # sort by km
        used = [False] * len(candidates)
        for i, (fnum_i, km_i, step_i) in enumerate(candidates):
            if used[i]:
                continue
            cluster = [(fnum_i, km_i, step_i)]
            for j in range(i + 1, len(candidates)):
                if used[j]:
                    continue
                _, km_j, _ = candidates[j]
                if abs(km_j - km_i) <= cluster_km:
                    cluster.append(candidates[j])
                    used[j] = True
            used[i] = True
            if len(cluster) >= min_fibers:
                for fnum_c, km_c, step_c in cluster:
                    validated[fnum_c].append({'km': km_c, 'step_db': step_c})

    return dict(validated)


# ─────────────────────────────────────────────────────────────────────────────
#  Break reflection measurement — classify break type from Fresnel height
# ─────────────────────────────────────────────────────────────────────────────

def measure_break_reflection(raw_db, dist, dx_km, pre_km, break_km):
    """
    Measure the Fresnel reflection height above the extrapolated Rayleigh
    baseline at a break position, and classify the break type.

    A fresh dig or aerial cut leaves a clean glass-air interface: the OTDR
    sees a strong Fresnel reflection spike above the Rayleigh line.
    A crushed or kinked fiber has no clean surface: the trace steps down
    without any spike — reflection height is near zero.

    Classification:
      CLEAN BREAK   — strong Fresnel (≥ REFL_CLEAN_BREAK dB)   → fresh dig / aerial cut
      PARTIAL BREAK — weak Fresnel  (≥ REFL_PARTIAL_BREAK dB)  → stress fracture / partial crush
      CRUSH/KINK    — faint Fresnel (≥ REFL_CRUSH_MIN dB)      → crush with small gap
      CRUSH         — no reflection (< REFL_CRUSH_MIN dB)       → severe crush / epoxy joint

    Returns (refl_height_db, classification_str) or (None, 'UNKNOWN').
    """
    i_break = int(round((break_km + pre_km) / dx_km))

    # Fit Rayleigh slope on 3 km of pre-break trace, stopping 0.5 km before break
    fit_end   = max(5, i_break - int(0.5 / dx_km))
    fit_start = max(5, fit_end - int(3.0 / dx_km))

    if fit_end - fit_start < 20 or i_break >= len(raw_db):
        return None, 'UNKNOWN'

    x_fit  = np.arange(fit_start, fit_end, dtype=float)
    y_fit  = raw_db[fit_start:fit_end]
    coeffs = np.polyfit(x_fit, y_fit, 1)

    # Search ±0.3 km around the detected break position for a positive spike
    srch_s = max(0, i_break - int(0.3 / dx_km))
    srch_e = min(len(raw_db) - 1, i_break + int(0.3 / dx_km))
    if srch_s >= srch_e:
        return None, 'UNKNOWN'

    x_srch   = np.arange(srch_s, srch_e, dtype=float)
    y_extrap = np.polyval(coeffs, x_srch)
    y_actual = raw_db[srch_s:srch_e]
    deviation = y_actual - y_extrap   # positive = above baseline = Fresnel spike

    refl_height = float(np.max(deviation))

    if refl_height >= REFL_CLEAN_BREAK:
        classification = 'CLEAN BREAK'
    elif refl_height >= REFL_PARTIAL_BREAK:
        classification = 'PARTIAL BREAK'
    elif refl_height >= REFL_CRUSH_MIN:
        classification = 'CRUSH/KINK'
    else:
        classification = 'CRUSH'

    return round(refl_height, 2), classification


# ─────────────────────────────────────────────────────────────────────────────
#  Bend zone detection — ribbon consensus
# ─────────────────────────────────────────────────────────────────────────────


def scan_trace_for_bends(r, known_event_kms=None, window_km=1.0,
                         step_km=0.25, guard_km=0.3):
    """
    Scan the RawSamples trace with a sliding window to find localized bends.

    Instead of one slope per inter-splice segment, this slides a window across
    the full trace and measures the local attenuation slope.  Positions where
    the local slope exceeds the fiber's baseline by > 0 are returned as
    candidate bend locations (the caller applies threshold + ribbon consensus).

    Parameters
    ----------
    r               : fiber dict with 'raw', 'dist', 'dx_km', 'noise_floor_km'
    known_event_kms : list of km positions to skip (splices, connectors)
    window_km       : width of the sliding fit window
    step_km         : step between consecutive windows
    guard_km        : distance around known events to exclude from fits

    Returns
    -------
    list of {km, local_slope, baseline_slope, excess} sorted by km
    """
    if not r.get('has_trace') or r.get('raw') is None:
        return []

    raw_db   = r['raw']
    dist     = r['dist']
    dx_km    = r['dx_km']
    noise_km = r['noise_floor_km']
    known    = known_event_kms or []

    # Build a mask that excludes ±guard_km around known events
    n = len(raw_db)
    ok = np.ones(n, dtype=bool)
    for ek in known:
        i_evt = int(round(ek / dx_km))
        i_guard = int(guard_km / dx_km)
        lo = max(0, i_evt - i_guard)
        hi = min(n, i_evt + i_guard)
        ok[lo:hi] = False

    # Don't use anything past the noise floor
    i_noise = int(min(noise_km, float(dist[-1])) / dx_km)
    ok[i_noise:] = False
    # Skip first 1 km (launch artifacts)
    i_start = int(1.0 / dx_km)
    ok[:i_start] = False

    # Global baseline slope (robust fit on all clean samples)
    clean_idx = np.where(ok)[0]
    if len(clean_idx) < 20:
        return []
    try:
        c_global = np.polyfit(dist[clean_idx], raw_db[clean_idx], 1)
        baseline_slope = abs(float(c_global[0]))
    except Exception:
        return []

    # Sliding window
    half_win = window_km / 2.0
    i_step   = max(1, int(step_km / dx_km))
    i_half   = int(half_win / dx_km)
    candidates = []

    i = i_start + i_half
    while i < i_noise - i_half:
        lo = i - i_half
        hi = i + i_half
        # Use only clean samples within the window
        win_ok = ok[lo:hi]
        if win_ok.sum() < 10:
            i += i_step
            continue
        win_idx = np.where(win_ok)[0] + lo
        try:
            c_local = np.polyfit(dist[win_idx], raw_db[win_idx], 1)
            local_slope = abs(float(c_local[0]))
        except Exception:
            i += i_step
            continue

        excess = local_slope - baseline_slope
        if excess > 0.005:   # noise floor — meaningful excess only
            candidates.append({
                'km':             round(float(dist[i]), 2),
                'local_slope':    round(local_slope, 5),
                'baseline_slope': round(baseline_slope, 5),
                'excess':         round(excess, 5),
            })
        i += i_step

    return candidates


def detect_trace_bend_zones(fibers_a, ribbon_size, bend_threshold,
                            known_event_kms=None, known_breaks=None,
                            fibers_b=None, span_km=None,
                            min_fibers_per_ribbon=3, cluster_km=0.8):
    """
    Detect bend zones from RawSamples trace shape using sliding window +
    ribbon consensus + optional bidirectional confirmation.

    When fibers_b and span_km are provided, the B-direction traces are also
    scanned.  A bend is symmetric (both directions see excess slope), so
    bidirectional agreement allows a LOWER threshold (half the normal) to
    catch subtle bends that either direction alone would miss — especially
    near the far end of the A trace where it's hitting the noise floor but
    the B trace has pristine signal.

    Returns:
      fiber_bends  — {fnum: [{km, excess}]}
      bend_columns — [{position_km, count, is_bend=True}] for positions not
                     already in the event-table-derived splice list
    """
    known_breaks = known_breaks or {}
    known_evts   = known_event_kms or []
    # B-direction known event positions (in B-frame for scanning)
    b_known_evts = [span_km - ek for ek in known_evts] if span_km else []

    # ── Pass 1: scan A-direction traces ──────────────────────────────────────
    # Store ALL candidates (even below threshold) for bidi cross-check
    a_cands_by_fiber = {}  # fnum → [{km, excess}]
    for fnum, r in fibers_a.items():
        break_km = known_breaks.get(fnum)
        cands = scan_trace_for_bends(r, known_event_kms=known_evts)
        if break_km is not None:
            cands = [c for c in cands if c['km'] < break_km]
        a_cands_by_fiber[fnum] = cands

    # ── Pass 2: scan B-direction traces, convert to A-frame ─────────────────
    b_cands_by_fiber = {}  # fnum → [{km (A-frame), excess}]
    if fibers_b and span_km:
        for fnum, r in fibers_b.items():
            cands = scan_trace_for_bends(r, known_event_kms=b_known_evts)
            # Convert B-frame km to A-frame km
            a_frame = [{'km': round(span_km - c['km'], 2),
                        'excess': c['excess']} for c in cands]
            b_cands_by_fiber[fnum] = a_frame

    # ── Pass 3: bidirectional confirmation ───────────────────────────────────
    # A bend is symmetric: both directions see excess slope at the same km.
    # If A sees excess > thresh/2 AND B sees excess > thresh/2 at the same
    # position (±0.5 km), confirm as a bend at the lower bidi threshold.
    # This catches subtle bends near the span ends where one direction has
    # poor SNR but the other has pristine signal.
    bidi_thresh = bend_threshold * 0.5  # half threshold when both agree

    ribbon_cands = defaultdict(list)
    for fnum in fibers_a:
        a_cands = a_cands_by_fiber.get(fnum, [])
        b_cands = b_cands_by_fiber.get(fnum, [])
        ri = (fnum - 1) // ribbon_size
        break_km = known_breaks.get(fnum)

        for c in a_cands:
            if break_km and c['km'] >= break_km:
                continue
            if c['excess'] >= bend_threshold:
                # A alone exceeds full threshold — always include
                ribbon_cands[ri].append((fnum, c['km'], c['excess']))
            elif c['excess'] >= bidi_thresh and b_cands:
                # A is borderline — check if B confirms
                b_match = [bc for bc in b_cands
                           if abs(bc['km'] - c['km']) < 0.5
                           and bc['excess'] >= bidi_thresh]
                if b_match:
                    # Bidirectional confirmation — use average excess
                    avg_ex = (c['excess'] + b_match[0]['excess']) / 2.0
                    ribbon_cands[ri].append((fnum, c['km'], avg_ex))

        # Also check B-only candidates (A noise floor may have missed them)
        if b_cands:
            for bc in b_cands:
                if break_km and bc['km'] >= break_km:
                    continue
                if bc['excess'] >= bend_threshold:
                    # B alone exceeds threshold — check if A has ANY signal
                    a_match = [ac for ac in a_cands
                               if abs(ac['km'] - bc['km']) < 0.5]
                    if a_match and a_match[0]['excess'] >= bidi_thresh:
                        avg_ex = (bc['excess'] + a_match[0]['excess']) / 2.0
                        ribbon_cands[ri].append((fnum, bc['km'], avg_ex))
                    elif not a_match:
                        # A has no data here (past noise floor) — trust B alone
                        # but require higher threshold to compensate
                        if bc['excess'] >= bend_threshold:
                            ribbon_cands[ri].append((fnum, bc['km'], bc['excess']))

    # Deduplicate per-fiber candidates (same fiber, same ~km)
    for ri in ribbon_cands:
        seen = {}
        deduped = []
        for fnum, km, excess in ribbon_cands[ri]:
            key = (fnum, round(km, 1))
            if key not in seen or excess > seen[key]:
                seen[key] = excess
                deduped.append((fnum, km, excess))
        ribbon_cands[ri] = deduped

    # ── Pass 4: ribbon consensus ─────────────────────────────────────────────
    fiber_bends = defaultdict(list)
    bend_positions = defaultdict(int)

    for ri, cands in ribbon_cands.items():
        if len(cands) < min_fibers_per_ribbon:
            continue
        cands.sort(key=lambda x: x[1])
        used = [False] * len(cands)
        for i in range(len(cands)):
            if used[i]:
                continue
            cluster = [cands[i]]
            for j in range(i + 1, len(cands)):
                if used[j]:
                    continue
                if abs(cands[j][1] - cands[i][1]) <= cluster_km:
                    cluster.append(cands[j])
                    used[j] = True
            used[i] = True
            if len(cluster) >= min_fibers_per_ribbon:
                avg_km = round(np.mean([c[1] for c in cluster]), 2)
                for fnum_c, km_c, excess_c in cluster:
                    fiber_bends[fnum_c].append({
                        'km':     km_c,
                        'excess': excess_c,
                    })
                    bend_positions[round(avg_km, 1)] += 1

    # Build candidate bend columns from consensus positions
    bend_columns = []
    for km_bin in sorted(bend_positions.keys()):
        bend_columns.append({
            'position_km': km_bin,
            'count':       bend_positions[km_bin],
            'is_bend':     True,
            'splice_num':  None,
        })

    return dict(fiber_bends), bend_columns


def _fiber_segment_slopes(r, splice_kms):
    """
    Measure attenuation slope in each inter-splice segment for one fiber.
    Returns list of {seg_idx, start_km, end_km, slope_db_km} or [].
    """
    if not r['has_trace']:
        return []

    dist        = r['dist']
    raw_db      = r['raw']
    noise_km    = r['noise_floor_km']
    boundaries  = [0.5] + sorted(splice_kms) + [min(noise_km, float(dist[-1])) * 0.98]
    segments    = []

    for seg_idx in range(len(boundaries) - 1):
        s_km = boundaries[seg_idx]     + 0.3
        e_km = boundaries[seg_idx + 1] - 0.3
        if e_km - s_km < 0.5:
            continue
        mask = (dist >= s_km) & (dist <= e_km)
        if mask.sum() < 10:
            continue
        try:
            coeffs = np.polyfit(dist[mask], raw_db[mask], 1)
            segments.append({
                'seg_idx':     seg_idx,
                'start_km':    s_km,
                'end_km':      e_km,
                'slope_db_km': abs(float(coeffs[0])),
            })
        except Exception:
            pass

    return segments


def detect_bend_zones(fibers_a, splices, ribbon_size, bend_threshold,
                      known_breaks=None, min_fibers_per_ribbon=3):
    """
    Detect cable bend zones using ribbon consensus.

    A bend in the cable bends ALL ribbons passing through it.  Within each
    ribbon, geometry means fibers on the outside radius suffer the most —
    so a bend shows up as a CLUSTER of fibers in the same ribbon with
    elevated attenuation in the same inter-splice segment.

    An isolated high-slope fiber (bad coating, local stress) won't pass the
    consensus test — it needs at least min_fibers_per_ribbon companions in
    the same ribbon and same segment.

    Returns {fnum: [{start_km, end_km, excess_db_km}]}
    """
    splice_kms = [sp['position_km'] for sp in splices]

    # Collect per-fiber slope data
    # seg_idx -> ribbon_idx -> [(fnum, excess, start_km, end_km)]
    seg_ribbon = defaultdict(lambda: defaultdict(list))

    known_breaks = known_breaks or {}

    for fnum, r in fibers_a.items():
        segs = _fiber_segment_slopes(r, splice_kms)
        if not segs:
            continue

        # If this fiber has a break, discard every segment past the break —
        # that portion of the trace is dead and produces meaningless slopes.
        break_km = known_breaks.get(fnum)
        if break_km is not None:
            segs = [s for s in segs if s['end_km'] <= break_km]
        if not segs:
            continue

        # Use this fiber's median slope as its personal baseline
        # (accounts for fiber-to-fiber variation in attenuation coefficient)
        slopes       = [s['slope_db_km'] for s in segs]
        fiber_median = float(np.median(slopes))
        ribbon_idx   = (fnum - 1) // ribbon_size

        for s in segs:
            excess = s['slope_db_km'] - fiber_median
            if excess > bend_threshold:
                seg_ribbon[s['seg_idx']][ribbon_idx].append(
                    (fnum, excess, s['start_km'], s['end_km']))

    # Apply consensus: only keep segments where ≥ min_fibers_per_ribbon agree
    fiber_bends = defaultdict(list)

    for seg_idx, ribbon_data in seg_ribbon.items():
        for ribbon_idx, fiber_list in ribbon_data.items():
            if len(fiber_list) >= min_fibers_per_ribbon:
                for fnum, excess, start_km, end_km in fiber_list:
                    fiber_bends[fnum].append({
                        'start_km':    round(start_km, 2),
                        'end_km':      round(end_km, 2),
                        'excess_db_km': round(excess, 3),
                    })

    return dict(fiber_bends)


# ─────────────────────────────────────────────────────────────────────────────
#  Full fiber loading
# ─────────────────────────────────────────────────────────────────────────────

def load_fiber(filepath):
    """
    Load a SOR file: events from KeyEvents + RawSamples trace.
    Returns dict or None.
    """
    try:
        with open(filepath, 'rb') as f:
            data = f.read()
    except OSError:
        return None

    blocks = _parse_block_directory(data)
    events = _parse_key_events(data, blocks)
    if not events:
        return None

    rs = read_rawsamples(filepath)
    if rs is None:
        # Fallback: return events only (no trace analysis)
        return {'events': events, 'has_trace': False,
                'raw': None, 'dist': None, 'dx_km': None,
                'pre_km': None, 'IOR': None, 'noise_floor_km': None}

    raw, dist, dx_km, pre_km, IOR = rs
    noise_km = find_noise_floor_km(dist, raw)

    return {
        'events':        events,
        'has_trace':     True,
        'raw':           raw,
        'dist':          dist,
        'dx_km':         dx_km,
        'pre_km':        pre_km,
        'IOR':           IOR,
        'noise_floor_km': noise_km,
    }


def load_all_fibers(dir_a, dir_b):
    fibers_a, fibers_b = {}, {}
    total = 0
    for d, store in [(dir_a, fibers_a), (dir_b, fibers_b)]:
        if not os.path.isdir(d):
            continue
        files = sorted(f for f in os.listdir(d) if f.lower().endswith('.sor'))
        for fname in files:
            digits = ''.join(
                c for c in fname.replace('_1550.sor', '').replace('.sor', '')
                if c.isdigit()
            )
            if not digits:
                continue
            fnum = int(digits)
            fpath = os.path.join(d, fname)
            r = load_fiber(fpath)
            if r:
                r['filepath'] = fpath
                store[fnum] = r
                total += 1
    return fibers_a, fibers_b


# ─────────────────────────────────────────────────────────────────────────────
#  Splice discovery
# ─────────────────────────────────────────────────────────────────────────────

MIN_BEND_POP = 5   # minimum fibers to create a bend column (tech shows zones with 2-7 events)

def discover_splices(fibers_a, n_fibers=None):
    """
    Discover splice/event positions from the A-direction event tables.

    Each position is classified as either a real splice closure (hardware
    splice that every fiber passes through → events in nearly all fibers)
    or a bend zone (distributed bend loss that affects only a subset of
    ribbons → fewer fibers show discrete events there).

    Classification: positions with events in >= 45% of fibers are real
    splices.  Positions with >= MIN_BEND_POP but < 45% are bend zones.

    Uses 0.1 km binning to resolve nearby but distinct positions (e.g.,
    65.03 km and 65.90 km) that 1-km binning would merge.
    """
    _nf = n_fibers or len(fibers_a)
    bins     = defaultdict(list)
    bins_tot = defaultdict(list)

    for fnum, r in fibers_a.items():
        for e in r['events']:
            if e['dist_km'] < 1.0 or e['is_end']:
                continue
            bk = round(e['dist_km'], 1)          # 0.1 km bins (was 1 km)
            bins[bk].append(e['dist_km'])
            bins_tot[bk].append(e['time_of_travel'])

    first = next(iter(fibers_a.values()))
    IOR   = 1.47
    for e in first['events']:
        if e['dist_km'] > 1 and e['time_of_travel'] > 0:
            IOR = e['time_of_travel'] * 0.02998 / (e['dist_km'] * 1000)
            break

    # Build candidate positions (lower threshold to catch small bend zones)
    splices = []
    for bk in sorted(bins.keys()):
        if len(bins[bk]) < MIN_BEND_POP:
            continue
        mode_tot = Counter(bins_tot[bk]).most_common(1)[0][0]
        mode_km  = round((mode_tot * 0.02998 / IOR) / 1000.0, 2)
        splices.append({'bin': bk, 'position_km': mode_km, 'count': len(bins[bk])})

    real_splice_min = max(MIN_POP_SPLICE, int(_nf * 0.45))

    # Pass 1: consolidate adjacent bins (< 0.15 km apart) into one position.
    # This merges bins split by the 0.1 km granularity (e.g., 8.0 + 8.1 = same splice).
    consolidated = []
    for sp in splices:
        if consolidated and abs(sp['position_km'] - consolidated[-1]['position_km']) < 0.15:
            # Merge into existing: keep position of the higher-count bin
            prev = consolidated[-1]
            prev['count'] += sp['count']
            if sp['count'] > prev.get('_primary_count', prev['count']):
                prev['position_km'] = sp['position_km']
                prev['bin'] = sp['bin']
            prev['_primary_count'] = max(prev.get('_primary_count', 0), sp['count'])
        else:
            sp['_primary_count'] = sp['count']
            consolidated.append(sp)

    # Pass 2: absorb small satellites (< 0.4 km) into confirmed real splices
    absorbed = set()
    for i, sp in enumerate(consolidated):
        if sp['count'] >= real_splice_min:
            for j, other in enumerate(consolidated):
                if j == i or j in absorbed:
                    continue
                if (abs(other['position_km'] - sp['position_km']) < 0.4
                        and other['count'] < real_splice_min):
                    sp['count'] += other['count']
                    absorbed.add(j)
    remaining = [sp for i, sp in enumerate(consolidated) if i not in absorbed]

    # Pass 3: merge remaining nearby entries, but never merge splice + bend
    merged = []
    for sp in remaining:
        if not merged:
            merged.append(sp)
            continue
        prev = merged[-1]
        gap = abs(sp['position_km'] - prev['position_km'])
        prev_is_real = (prev['count'] >= real_splice_min)
        curr_is_real = (sp['count'] >= real_splice_min)

        if gap < 0.6:
            if prev_is_real != curr_is_real:
                merged.append(sp)          # splice + bend → keep both
            elif sp['count'] > prev['count']:
                merged[-1] = sp            # same type → keep stronger
        else:
            merged.append(sp)

    # Sort by position (merging can reorder)
    merged.sort(key=lambda s: s['position_km'])

    # Classify: real splice vs. bend zone by fiber coverage
    splice_num = 0
    for sp in merged:
        sp['is_bend'] = (sp['count'] < real_splice_min)
        if not sp['is_bend']:
            splice_num += 1
            sp['splice_num'] = splice_num
        else:
            sp['splice_num'] = None

    return merged


# ─────────────────────────────────────────────────────────────────────────────
#  Per-fiber analysis
# ─────────────────────────────────────────────────────────────────────────────

def analyze_fiber(fnum, ra, rb, splices, span_km, threshold,
                  precomputed_break_km=None):
    """
    Full analysis for one fiber in both directions.

    precomputed_break_km: rough break position from ribbon EOL comparison.
    If supplied, we skip internal detection and go straight to trace localization.

    Returns:
      splice_results  — [{splice_idx, loss, source, flag, is_break, is_bend, a_blind}]
      break_info      — {'break_km', 'step_loss_db', 'offset_ft', 'loss', 'type'} or None
      bend_segments   — [{start_km, end_km, excess_db_km}]
      noise_a_km      — noise floor from A direction
    """
    splice_kms = [sp['position_km'] for sp in splices]

    # ── Noise floor ────────────────────────────────────────────────────────
    noise_a_km = span_km
    if ra and ra['has_trace']:
        noise_a_km = ra['noise_floor_km']

    # ── Break localization ─────────────────────────────────────────────────
    # Detection was already done (ribbon EOL comparison) — here we only
    # refine the position using the raw trace step-loss method.
    break_a = None

    if precomputed_break_km is not None:
        rough_km   = precomputed_break_km
        precise_km = rough_km
        step_loss  = None

        refl_height    = None
        break_class    = 'UNKNOWN'

        if ra and ra['has_trace']:
            precise_km, step_loss = localize_break_step(
                ra['raw'], ra['dist'], ra['dx_km'], ra['pre_km'], rough_km)
            refl_height, break_class = measure_break_reflection(
                ra['raw'], ra['dist'], ra['dx_km'], ra['pre_km'], precise_km)

        # Offset from nearest splice
        nearest_splice_km = None
        if splice_kms:
            nearest_splice_km = min(splice_kms, key=lambda sk: abs(sk - precise_km))
        offset_ft = (precise_km - nearest_splice_km) * 3280.84 \
                    if nearest_splice_km is not None else None

        # Get splice loss from event table at break position (if recorded)
        brk_loss = 0.0
        if ra:
            for e in ra['events']:
                if abs(e['dist_km'] - precise_km) < 1.0 and not e['is_end']:
                    brk_loss = e['splice_loss']
                    break

        break_a = {
            'break_km':    precise_km,
            'rough_km':    rough_km,
            'step_loss':   step_loss,
            'offset_ft':   round(offset_ft, 0) if offset_ft is not None else None,
            'loss':        brk_loss,
            'refl_height': refl_height,
            'break_class': break_class,
            'type':        'BREAK' if step_loss is not None else 'broke',
        }

    # ── B-direction eof ────────────────────────────────────────────────────
    eof_b = span_km
    if rb:
        for e in rb['events']:
            if e['is_end']:
                eof_b = e['dist_km']
                break

    # ── Pre-compute known event positions for window trimming ──────────────
    known_kms_a = [e['dist_km'] for e in ra['events'] if not e['is_end']] if ra else []
    known_kms_b = []
    if rb:
        known_kms_b = [eof_b - e['dist_km']
                       for e in rb['events'] if not e['is_end']]

    # ── Scan for events the OTDR event table missed (A direction only) ──────
    missed_a = []
    if ra and ra['has_trace']:
        missed_a = scan_for_missed_events(
            ra['raw'], ra['dist'], ra['dx_km'], ra['pre_km'],
            noise_a_km, known_kms_a)

    # ── Per-splice results ─────────────────────────────────────────────────
    splice_results = []

    for si, sp in enumerate(splices):
        skm    = sp['position_km']
        loss_a = None
        loss_b = None
        loss_a_trace = None   # direct trace measurement
        loss_b_trace = None
        pre_slope_a  = None   # attenuation slope approaching splice from A
        post_slope_a = None   # attenuation slope leaving splice toward B

        # A-direction blind past break?
        a_blind = (break_a is not None and skm > break_a['break_km'] + 0.5)

        # ── A-direction: try raw trace first, fall back to event table ──────
        if ra and not a_blind:
            if ra['has_trace']:
                loss_a_trace, pre_slope_a, post_slope_a = measure_event_loss_from_trace(
                    ra['raw'], ra['dist'], ra['dx_km'], ra['pre_km'],
                    skm,
                    noise_km=ra['noise_floor_km'],
                    other_event_kms=[k for k in known_kms_a if abs(k - skm) > 0.1])

            # Event table fallback (also used for validation)
            loss_a_evt = None
            for e in ra['events']:
                if (abs(e['dist_km'] - skm) < SPLICE_MATCH_KM
                        and e['dist_km'] > 0.5 and not e['is_end']):
                    loss_a_evt = e['splice_loss']
                    break

            # Use trace measurement if valid (positive and not unreasonably large)
            # Fall back to event table if trace measurement failed or looks wrong
            if loss_a_trace is not None and 0 <= loss_a_trace <= 5.0:
                loss_a = loss_a_trace
            elif loss_a_evt is not None:
                loss_a = loss_a_evt

        # ── B-direction: same approach, convert to A-frame ──────────────────
        if rb:
            b_km_pos = eof_b - skm   # position in B-direction frame
            if rb.get('has_trace') and rb.get('raw') is not None:
                b_noise_km = rb.get('noise_floor_km', eof_b)
                loss_b_trace, _, _ = measure_event_loss_from_trace(
                    rb['raw'], rb['dist'], rb['dx_km'], rb['pre_km'],
                    b_km_pos,
                    noise_km=b_noise_km,
                    other_event_kms=[eof_b - k for k in known_kms_b
                                     if abs((eof_b - k) - b_km_pos) > 0.1])

            loss_b_evt = None
            for e in rb['events']:
                e_from_a = eof_b - e['dist_km']
                if (abs(e_from_a - skm) < SPLICE_MATCH_KM
                        and e['dist_km'] > 0.5 and not e['is_end']):
                    loss_b_evt = e['splice_loss']
                    break

            if loss_b_trace is not None and 0 <= loss_b_trace <= 5.0:
                loss_b = loss_b_trace
            elif loss_b_evt is not None:
                loss_b = loss_b_evt

        # Determine reporting source
        if not a_blind and loss_a is not None:
            primary = loss_a
            source  = 'A'
        elif a_blind and loss_b is not None:
            primary = loss_b
            source  = 'B_fill'
        elif not a_blind and loss_a is None and loss_b is not None:
            primary = loss_b
            source  = 'B_fill'
        else:
            primary = None
            source  = 'A' if not a_blind else 'B_fill'

        # Is this the break splice?
        is_break = (
            break_a is not None
            and abs(skm - break_a['break_km']) < SPLICE_MATCH_KM
        )

        # Bend at this splice?
        is_bend = (
            loss_a is not None and loss_b is not None
            and loss_a > 0.05 and loss_b > 0.05
            and abs(loss_a - loss_b) < 0.15
        )

        # Bidirectional average — truest splice loss (removes directionality)
        bidi_loss = None
        if loss_a is not None and loss_b is not None:
            bidi_loss = round((loss_a + loss_b) / 2.0, 4)

        flagged = (primary is not None and abs(primary) >= threshold)

        # Gainer/loser classification:
        # At a splice, mode-field mismatch makes one direction see gain (negative)
        # and the other see loss (positive).  True damage is always loss from
        # both sides — it never creates apparent gain.
        #   'both_loss' — both A and B show positive loss → suspicious for damage
        #   'gainer_loser' — one positive, one negative → normal mode-field mismatch
        #   None — insufficient bidirectional data
        gainer_loser = None
        if loss_a is not None and loss_b is not None:
            if loss_a > 0 and loss_b > 0:
                gainer_loser = 'both_loss'
            elif (loss_a > 0 and loss_b < 0) or (loss_a < 0 and loss_b > 0):
                gainer_loser = 'gainer_loser'

        splice_results.append({
            'splice_idx':  si,
            'loss_a':      loss_a,
            'loss_b':      loss_b,
            'bidi_loss':   bidi_loss,
            'primary':     primary,
            'source':      source,
            'flagged':     flagged,
            'is_break':    is_break,
            'is_bend':     is_bend,
            'a_blind':     a_blind,
            'pre_slope_a': pre_slope_a,
            'post_slope_a':post_slope_a,
            'gainer_loser': gainer_loser,
        })

    return splice_results, break_a, noise_a_km, missed_a


# ─────────────────────────────────────────────────────────────────────────────
#  Loss formatting helpers
# ─────────────────────────────────────────────────────────────────────────────

def fmt_loss(loss):
    """Format loss value in tech notation: .162 for <1 dB, 1.200 for >=1 dB."""
    if loss is None:
        return '?'
    if abs(loss) < 1.0:
        sign = '-' if loss < 0 else ''
        return f"{sign}.{abs(loss) * 1000:03.0f}"
    return f"{loss:.3f}"


def group_fibers(fiber_losses, threshold, bend_notation=False):
    """
    Given {fnum: loss_value} for fibers in a ribbon, group consecutive fibers
    with similar losses (within ±0.015 dB) into shorthand notation.

    Returns list of strings like:
      "37 .180"           — single fiber
      "49-56 .220"        — range with similar losses
      "49-56 .220's"      — range in a bend column (bend_notation=True)
    """
    if not fiber_losses:
        return []

    items  = sorted(fiber_losses.items())
    groups = []
    i      = 0
    while i < len(items):
        fnum, loss = items[i]
        # Try to extend a run of similar losses
        j = i + 1
        while j < len(items):
            fnext, lnext = items[j]
            if fnext == items[j - 1][0] + 1 and abs(lnext - loss) <= 0.015:
                j += 1
            else:
                break
        # items[i..j-1] form a group
        if j - i >= 3:
            avg_loss = float(np.mean([v for _, v in items[i:j]]))
            suffix = "'s" if bend_notation else ''
            groups.append(f"{items[i][0]}-{items[j-1][0]} {fmt_loss(avg_loss)}{suffix}")
        else:
            for k in range(i, j):
                fn, fl = items[k]
                groups.append(f"{fn} {fmt_loss(fl)}")
        i = j

    return groups


# ─────────────────────────────────────────────────────────────────────────────
#  Excel output
# ─────────────────────────────────────────────────────────────────────────────

def write_xlsx(all_results, all_breaks, all_bends, all_missed, splices, n_fibers,
               output_path, site_a, site_b, span_km, threshold, ribbon_size):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # ── Sheet 1: Damage Report ────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Damage Report'

    # Styles (matching splice_report.xlsx exactly)
    hdr_fill     = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    bend_hdr_fill= PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    hdr_font     = Font(bold=True, size=10, color='FFFFFF')
    bend_hdr_font= Font(bold=True, size=10, color='7B6600')
    a_km_font    = Font(bold=True, size=9, color='1F4E79')
    a_ft_font    = Font(bold=True, size=9, color='2E75B6')
    b_km_font    = Font(bold=True, size=9, color='833C00')
    b_ft_font    = Font(bold=True, size=9, color='C55A11')
    rib_font     = Font(size=9)
    cell_font    = Font(size=8)
    brk_font     = Font(size=8, bold=True, color='FFFFFF')
    bfill_font   = Font(size=8, color='0070C0')

    pink_fill   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    red_fill    = PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid')
    orange_fill = PatternFill(start_color='FF8800', end_color='FF8800', fill_type='solid')
    blue_fill   = PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid')
    gold_fill   = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    bend_fill   = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    n_cols = len(splices)
    # Layout: col 1 = Ribbon, col 2 = ILA:A, cols 3..3+n_cols-1 = splice/bend,
    #         col 3+n_cols = ILA:B
    ila_b_col = n_cols + 3

    # ── Pre-compute per-bend-column stats for header labels ──────────────────
    # Scan all_results to find max loss at each bend position, and count how
    # many ribbons are affected (for odd/even pattern detection).
    bend_col_stats = {}   # si → {'max_loss', 'event_count', 'affected_ribbons'}
    for si, sp in enumerate(splices):
        if not sp['is_bend']:
            continue
        max_loss = 0.0
        evt_count = 0
        affected_ribbons = set()
        for fnum_si, rec in all_results.items():
            fnum, s_idx = fnum_si
            if s_idx != si:
                continue
            p = rec.get('primary')
            if p is not None and p > 0.03:    # above noise floor
                max_loss = max(max_loss, p)
                evt_count += 1
                ri = (fnum - 1) // ribbon_size
                affected_ribbons.add(ri)
        # If no data from all_results, use event table count
        if evt_count == 0:
            evt_count = sp['count']
        # Detect odd/even ribbon pattern (helix geometry)
        ribbon_pattern = None
        if len(affected_ribbons) >= 3:
            sides = [ri % 2 for ri in affected_ribbons]
            if all(s == 0 for s in sides):
                ribbon_pattern = 'odd ribbons'    # 0-indexed even = tube side 1
            elif all(s == 1 for s in sides):
                ribbon_pattern = 'even ribbons'   # 0-indexed odd = tube side 2

        bend_col_stats[si] = {
            'max_loss': max_loss,
            'event_count': evt_count,
            'affected_ribbons': sorted(affected_ribbons),
            'is_sub_threshold': (max_loss < threshold and max_loss > 0),
            'ribbon_pattern': ribbon_pattern,
        }

    # ── Header rows 1-2: distances (km / ft, both directions, condensed) ──────
    # Row 1: A→B — "8.05km / 26,411ft"
    # Row 2: B→A — "81.95km / 268,845ft"
    ws.cell(row=1, column=2, value='A→B:').font = a_km_font
    ws.cell(row=2, column=2, value='B→A:').font = b_km_font

    for si, sp in enumerate(splices):
        col  = si + 3
        km   = sp['position_km']
        ft   = km * 3280.84
        b_km = span_km - km
        b_ft = b_km * 3280.84
        c1 = ws.cell(row=1, column=col, value=f"{km:.2f}km / {ft:,.0f}ft")
        c1.font      = a_km_font
        c1.alignment = Alignment(horizontal='center')
        c2 = ws.cell(row=2, column=col, value=f"{b_km:.2f}km / {b_ft:,.0f}ft")
        c2.font      = b_km_font
        c2.alignment = Alignment(horizontal='center')

    c1 = ws.cell(row=1, column=ila_b_col,
                 value=f"{span_km:.2f}km / {span_km*3280.84:,.0f}ft")
    c1.font = a_km_font
    c2 = ws.cell(row=2, column=ila_b_col, value='0.00km / 0ft')
    c2.font = b_km_font

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    HDR_ROW = 3
    ws.cell(row=HDR_ROW, column=1, value='Ribbon').font = hdr_font
    ws.cell(row=HDR_ROW, column=1).fill                 = hdr_fill
    ws.cell(row=HDR_ROW, column=2, value=f'ILA:{site_a}').font = hdr_font
    ws.cell(row=HDR_ROW, column=2).fill                 = hdr_fill

    for si, sp in enumerate(splices):
        col = si + 3
        if sp['is_bend']:
            stats = bend_col_stats.get(si, {})
            ml = stats.get('max_loss', 0)
            if ml > 0 and ml < threshold:
                # Sub-threshold bend — label like tech: "Bends under .08"
                label = f"Bends under {fmt_loss(ml)}"
            else:
                label = "Bends"
            c = ws.cell(row=HDR_ROW, column=col, value=label)
            c.font      = bend_hdr_font
            c.fill      = bend_hdr_fill
            c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        else:
            label = f"Splice {sp['splice_num']}"
            c = ws.cell(row=HDR_ROW, column=col, value=label)
            c.font = hdr_font
            c.fill = hdr_fill

    c = ws.cell(row=HDR_ROW, column=ila_b_col, value=f'ILA:{site_b}')
    c.font = hdr_font
    c.fill = hdr_fill

    # Tail column — shows EOL distance for short/broken fibers
    tail_col = ila_b_col + 1
    c = ws.cell(row=HDR_ROW, column=tail_col, value='Tail:')
    c.font = hdr_font
    c.fill = hdr_fill

    # Tube codes (A1-X2 for 864 fiber cable, A1-R2 for 432)
    def tube_code(ribbon_idx, n_fibers):
        letters = 'ABCDEFGHIJKLMNOPQRSTUVWX' if n_fibers > 432 else 'ABCDEFGHIJKLMNOPQR'
        letter_idx = ribbon_idx // 2
        side       = (ribbon_idx % 2) + 1
        return f'({letters[letter_idx]}{side})' if letter_idx < len(letters) else ''

    # ── Data rows (row 6+) ────────────────────────────────────────────────────
    DATA_START = HDR_ROW + 1
    n_ribbons  = (n_fibers + ribbon_size - 1) // ribbon_size

    for ri in range(n_ribbons):
        row     = DATA_START + ri
        f_start = ri * ribbon_size + 1
        f_end   = min(f_start + ribbon_size - 1, n_fibers)
        tc      = tube_code(ri, n_fibers)
        label   = f'Fiber {f_start}-{f_end} ({ri+1}) {tc}'.strip()
        ws.cell(row=row, column=1, value=label).font = rib_font

        for si, sp in enumerate(splices):
            col       = si + 3
            is_bend_col = sp['is_bend']

            # Collect all fibers in this ribbon at this position
            normal_a_losses = {}
            flagged_losses  = {}   # flagged, B also confirmed
            flagged_uni     = {}   # flagged, A only (no B confirmation)
            break_entries   = []
            bfill_losses    = {}
            has_broke       = []
            bend_fibers     = {}
            cell_has_break  = False
            cell_has_broke  = False

            for fnum in range(f_start, f_end + 1):
                rec = all_results.get((fnum, si))
                if rec is None:
                    continue

                brk = all_breaks.get(fnum)

                if rec['is_break'] and brk is not None:
                    loss_str   = fmt_loss(brk['loss'])
                    brk_cls    = brk.get('break_class', '')
                    cls_abbrev = {
                        'CLEAN BREAK':   'CLEAN',
                        'PARTIAL BREAK': 'PARTIAL',
                        'CRUSH/KINK':    'CRUSH/KINK',
                        'CRUSH':         'CRUSH',
                        'UNKNOWN':       '',
                    }.get(brk_cls, brk_cls)
                    if brk['type'] == 'BREAK' and brk['offset_ft'] is not None:
                        off_ft = abs(brk['offset_ft'])
                        suffix = f' [{cls_abbrev}]' if cls_abbrev else ''
                        break_entries.append(
                            f"{fnum} BREAK {loss_str} ({off_ft:.0f}ft){suffix}")
                        cell_has_break = True
                    else:
                        cls_suffix = f' [{cls_abbrev}]' if cls_abbrev else ''
                        has_broke.append(f"{fnum}{cls_suffix}")
                        cell_has_broke = True
                    continue

                if rec['a_blind'] and rec['primary'] is None:
                    has_broke.append(fnum)
                    cell_has_broke = True
                    continue

                p = rec['primary']
                if p is None or p < 0:   # suppress negatives (Fresnel artifacts)
                    continue

                if rec['is_bend']:
                    bend_fibers[fnum] = p
                elif rec['source'] == 'B_fill':
                    if p > 0:
                        bfill_losses[fnum] = p
                elif rec['flagged']:
                    # Determine if B direction confirmed the flag
                    loss_b = rec.get('loss_b')
                    b_confirmed = (loss_b is not None and loss_b >= threshold)
                    if b_confirmed:
                        flagged_losses[fnum] = p
                    else:
                        flagged_uni[fnum] = p
                else:
                    if abs(p) >= threshold * 0.5:
                        normal_a_losses[fnum] = p

            # Build cell content
            parts     = []
            cell_fill = None

            if break_entries:
                parts.extend(break_entries)
                cell_fill = red_fill

            if has_broke:
                # has_broke entries are strings like "49 [CRUSH]" or plain "49"
                # Try range compression only when all entries are plain fiber numbers
                plain_fnums = []
                mixed_entries = []
                for entry in has_broke:
                    if ' ' not in str(entry):
                        plain_fnums.append(int(entry))
                    else:
                        mixed_entries.append(str(entry) + ' broke')
                if plain_fnums and not mixed_entries:
                    plain_fnums.sort()
                    if len(plain_fnums) > 1 and (plain_fnums[-1] - plain_fnums[0] == len(plain_fnums) - 1):
                        parts.append(f"{plain_fnums[0]}-{plain_fnums[-1]} broke")
                    else:
                        for fn in plain_fnums:
                            parts.append(f"{fn} broke")
                else:
                    for fn in plain_fnums:
                        parts.append(f"{fn} broke")
                    parts.extend(mixed_entries)
                if cell_fill is None:
                    cell_fill = orange_fill

            # .XX's notation on for all columns (improvement #4)
            if flagged_losses:
                for grp in group_fibers(flagged_losses, threshold, bend_notation=True):
                    parts.append(grp)
                if cell_fill is None:
                    cell_fill = bend_fill if is_bend_col else pink_fill

            if flagged_uni:
                for grp in group_fibers(flagged_uni, threshold, bend_notation=True):
                    parts.append(f"{grp} uni")
                if cell_fill is None:
                    cell_fill = bend_fill if is_bend_col else pink_fill

            if bend_fibers:
                for grp in group_fibers(bend_fibers, threshold, bend_notation=True):
                    parts.append(f"{grp} bend")
                if cell_fill is None:
                    cell_fill = gold_fill

            if bfill_losses:
                for grp in group_fibers(bfill_losses, threshold, bend_notation=True):
                    parts.append(f"{grp} (B)")
                if cell_fill is None:
                    cell_fill = blue_fill

            if normal_a_losses:
                for grp in group_fibers(normal_a_losses, threshold, bend_notation=True):
                    parts.append(grp)
                if cell_fill is None:
                    cell_fill = bend_fill if is_bend_col else pink_fill

            # Sub-threshold bend columns: show only "N events" in first ribbon
            # row and skip per-fiber detail (matches tech notation).
            stats = bend_col_stats.get(si, {})
            if is_bend_col and stats.get('is_sub_threshold'):
                if ri == 0:
                    evt_ct = stats.get('event_count', sp['count'])
                    parts = [f"{evt_ct} events"]
                    cell_fill = bend_fill
                else:
                    # Sub-threshold: suppress per-fiber detail in data rows
                    parts = []
                    cell_fill = bend_fill

            # Event count summary in first ribbon row of above-threshold bend cols
            elif is_bend_col and ri == 0 and sp['count'] > 0:
                parts.append(f"({sp['count']} events total)")

            # For bend columns: use yellow fill even if cell is empty
            # (shows the column is a bend position at a glance)
            if is_bend_col and cell_fill is None and not parts:
                cell_fill = bend_fill

            if parts:
                cell = ws.cell(row=row, column=col, value='\n'.join(parts))
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if cell_has_break:
                    cell.font = brk_font
                elif cell_has_broke:
                    cell.font = Font(size=8, bold=True, color='FFFFFF')
                elif bfill_losses and not flagged_losses and not normal_a_losses:
                    cell.font = bfill_font
                elif is_bend_col:
                    cell.font = Font(size=8, color='7B6600')
                else:
                    cell.font = cell_font
                if cell_fill:
                    cell.fill = cell_fill
            elif cell_fill:
                ws.cell(row=row, column=col).fill = cell_fill

        # Tail column — short/broken fibers' EOL distance
        tail_entries = []
        for fnum in range(f_start, f_end + 1):
            brk = all_breaks.get(fnum)
            if brk and brk.get('break_km'):
                bk = brk['break_km']
                ft = bk * 3280.84
                tail_entries.append(f"{fnum} {bk:.2f}km/{ft:,.0f}ft")
        if tail_entries:
            tc = ws.cell(row=row, column=tail_col,
                         value='\n'.join(tail_entries))
            tc.alignment = Alignment(wrap_text=True, vertical='top')
            tc.font = Font(size=8, color='CC0000')

        # Attenuation-slope data after tail column (excess dB/km notes)
        attn_note_col = tail_col + 1
        ribbon_attn = []
        for fnum in range(f_start, f_end + 1):
            segs = all_bends.get(fnum, [])
            for seg in segs:
                ribbon_attn.append(
                    f"{fnum} +{seg['excess_db_km']:.3f}dB/km "
                    f"({seg['start_km']:.1f}-{seg['end_km']:.1f}km)"
                )
        # Store attenuation notes
        if ribbon_attn:
            ac = ws.cell(row=row, column=attn_note_col,
                         value='\n'.join(ribbon_attn))
            ac.alignment = Alignment(wrap_text=True, vertical='top')
            ac.font      = Font(size=7, color='7B6600', italic=True)
            ac.fill      = PatternFill(start_color='FFFACD',
                                       end_color='FFFACD', fill_type='solid')

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    for si, sp in enumerate(splices):
        letter = openpyxl.utils.get_column_letter(si + 3)
        ws.column_dimensions[letter].width = 18 if sp['is_bend'] else 22
    ws.column_dimensions[openpyxl.utils.get_column_letter(ila_b_col)].width     = 10
    ws.column_dimensions[openpyxl.utils.get_column_letter(ila_b_col + 1)].width = 30

    # ── Sheet 2: Breaks ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Breaks')
    ws2['A1'] = f'Break Detail — {site_a} → {site_b}'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A2'] = f'Span: {span_km} km  ({span_km * 3280.84:,.0f} ft)  |  Threshold: {threshold} dB'

    hdrs = ['Fiber', 'Break km (A→)', 'Break ft (A→)',
            'Break km (←B)', 'Break ft (←B)',
            'Step loss (dB)', 'Offset from splice (ft)',
            'Loss at break (dB)', 'Fresnel refl. above baseline (dB)', 'Break Classification']
    for ci, h in enumerate(hdrs, 1):
        c = ws2.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = hdr_fill

    # Color coding for break classifications
    class_fill = {
        'CLEAN BREAK':   PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid'),
        'PARTIAL BREAK': PatternFill(start_color='FF8800', end_color='FF8800', fill_type='solid'),
        'CRUSH/KINK':    PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'CRUSH':         PatternFill(start_color='DDEEFF', end_color='DDEEFF', fill_type='solid'),
        'UNKNOWN':       pink_fill,
    }
    class_font = {
        'CLEAN BREAK':   Font(bold=True, size=10, color='FFFFFF'),
        'PARTIAL BREAK': Font(bold=True, size=10, color='FFFFFF'),
        'CRUSH/KINK':    Font(size=10),
        'CRUSH':         Font(size=10),
        'UNKNOWN':       Font(size=10),
    }

    splice_kms = [sp['position_km'] for sp in splices]

    row = 5
    for fnum in sorted(all_breaks.keys()):
        b   = all_breaks[fnum]
        bkm = b['break_km']
        bc  = b.get('break_class', 'UNKNOWN')
        rh  = b.get('refl_height')

        ws2.cell(row=row, column=1,  value=fnum)
        ws2.cell(row=row, column=2,  value=round(bkm, 3))
        ws2.cell(row=row, column=3,  value=round(bkm * 3280.84, 0))
        ws2.cell(row=row, column=4,  value=round(span_km - bkm, 3))
        ws2.cell(row=row, column=5,  value=round((span_km - bkm) * 3280.84, 0))
        ws2.cell(row=row, column=6,
                 value=round(b['step_loss'], 2) if b['step_loss'] else '—')
        ws2.cell(row=row, column=7,
                 value=round(b['offset_ft'], 0) if b['offset_ft'] is not None else '—')
        ws2.cell(row=row, column=8,  value=round(b['loss'], 3) if b['loss'] else '—')
        ws2.cell(row=row, column=9,  value=round(rh, 2) if rh is not None else '—')
        c10 = ws2.cell(row=row, column=10, value=bc)
        c10.fill = class_fill.get(bc, pink_fill)
        c10.font = class_font.get(bc, Font(size=10))

        for ci in range(1, 10):
            ws2.cell(row=row, column=ci).fill = class_fill.get(bc, pink_fill)
            if bc in ('CLEAN BREAK', 'PARTIAL BREAK'):
                ws2.cell(row=row, column=ci).font = Font(size=10, color='FFFFFF')

        row += 1

    # ── Sheet 3: Missed Events ────────────────────────────────────────────────
    ws_miss = wb.create_sheet('Missed Events')
    ws_miss['A1'] = f'OTDR-Missed Events — {site_a} → {site_b}'
    ws_miss['A1'].font = Font(bold=True, size=14)
    ws_miss['A2'] = (
        'Events found in the raw trace that the OTDR event table did not record. '
        'Measured by sliding-window step-loss scan on RawSamples.'
    )
    miss_hdrs = ['Fiber', 'Ribbon', 'Position km', 'Position ft', 'Step loss (dB)',
                 'Nearest known splice', 'Nearest splice km', 'Distance from splice (ft)']
    for ci, h in enumerate(miss_hdrs, 1):
        c = ws_miss.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = hdr_fill

    miss_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    splice_kms_list = [(i, sp['position_km'], sp) for i, sp in enumerate(splices)]

    mr = 5
    for fnum in sorted(all_missed.keys()):
        ribbon_idx = (fnum - 1) // ribbon_size
        for ev in all_missed[fnum]:
            ekm = ev['km']
            # Find nearest known splice
            nearest_si, nearest_km, nearest_sp = min(
                splice_kms_list, key=lambda t: abs(t[1] - ekm))
            near_label = (f"Splice {nearest_sp['splice_num']}"
                          if not nearest_sp['is_bend']
                          else f"Bends {nearest_km:.2f}km")
            dist_ft = (ekm - nearest_km) * 3280.84

            ws_miss.cell(row=mr, column=1, value=fnum)
            ws_miss.cell(row=mr, column=2, value=ribbon_idx + 1)
            ws_miss.cell(row=mr, column=3, value=round(ekm, 3))
            ws_miss.cell(row=mr, column=4, value=round(ekm * 3280.84, 0))
            ws_miss.cell(row=mr, column=5, value=ev['step_db'])
            ws_miss.cell(row=mr, column=6, value=near_label)
            ws_miss.cell(row=mr, column=7, value=round(nearest_km, 3))
            ws_miss.cell(row=mr, column=8, value=round(dist_ft, 0))
            for ci in range(1, 9):
                ws_miss.cell(row=mr, column=ci).fill = miss_fill
            mr += 1

    if mr == 5:
        ws_miss['A5'] = 'No missed events detected — OTDR event table appears complete.'

    # ── Sheet 4: Legend ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Legend')
    ws3['A1'] = 'TechReplacement — Damage Report Legend'
    ws3['A1'].font = Font(bold=True, size=14)

    legend_items = [
        ('FFC7CE', 'Pink',        f'Splice loss ≥ {threshold} dB (A-direction)'),
        ('FF4444', 'Red',         'BREAK — break located with Fresnel peak offset'),
        ('FF8800', 'Orange',      '"broke" — confirmed break, precise offset unavailable'),
        ('DDEEFF', 'Light blue',  'B-fill — A was blind here; B-direction data used'),
        ('FFD966', 'Gold',        'Bend loss — symmetric in both directions (cable bend)'),
        ('FFFF00', 'Yellow',      'Bends column — elevated attenuation segment (dB/km excess)'),
    ]
    for i, (color, name, desc) in enumerate(legend_items, 4):
        ws3.cell(row=i, column=1, value='  ').fill = \
            PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws3.cell(row=i, column=2, value=name).font = Font(bold=True)
        ws3.cell(row=i, column=3, value=desc)

    ws3['A11'] = 'Cell format:  {fiber} {.loss}  |  BREAK: {fiber} BREAK {loss} ({offset}ft from splice)'
    ws3['A12'] = 'B-fill:       {fiber} {.loss} (B)  — data from B-direction trace'
    ws3['A13'] = f'Threshold: {threshold} dB  |  Span: {site_a}→{site_b}  {span_km} km'

    wb.save(output_path)
    print(f"  Saved: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description='TechReplacement — Automated OTDR Damage Report')
    ap.add_argument('dir_a',  help='A-direction SOR folder')
    ap.add_argument('dir_b',  help='B-direction SOR folder')
    ap.add_argument('--threshold',      type=float, default=0.15,
                    help='Unidir splice loss fail threshold dB (EXFO: 0.15)')
    ap.add_argument('--bidi-threshold', type=float, default=0.30,
                    help='Bidir splice loss fail threshold dB (EXFO: 0.30)')
    ap.add_argument('--bend-threshold', type=float, default=0.05,
                    help='Bend excess attenuation dB/km (default 0.05)')
    ap.add_argument('--output',    default='tech_damage_report.xlsx',
                    help='Output Excel file')
    ap.add_argument('--site-a',    default='A',  help='Site A name')
    ap.add_argument('--site-b',    default='B',  help='Site B name')
    ap.add_argument('--ribbon-size', type=int, default=12,
                    help='Fibers per ribbon (default 12)')
    args = ap.parse_args()

    print('TechReplacement — Automated OTDR Damage Report')
    print(f'  A: {args.dir_a}')
    print(f'  B: {args.dir_b}')
    print(f'  Unidir threshold: {args.threshold} dB  |  Bidir threshold: {args.bidi_threshold} dB'
          f'  |  Bend threshold: {args.bend_threshold} dB/km')

    print('\nLoading SOR files (event table + RawSamples)...')
    fibers_a, fibers_b = load_all_fibers(args.dir_a, args.dir_b)
    n_fibers = max(max(fibers_a.keys(), default=0),
                   max(fibers_b.keys(), default=0))
    trace_a  = sum(1 for r in fibers_a.values() if r['has_trace'])
    trace_b  = sum(1 for r in fibers_b.values() if r['has_trace'])
    print(f'  A: {len(fibers_a)} fibers ({trace_a} with RawSamples)')
    print(f'  B: {len(fibers_b)} fibers ({trace_b} with RawSamples)')

    print('\nDiscovering splice positions...')
    if not fibers_a:
        print('  ERROR: No A-direction SOR files found.')
        sys.exit(1)
    splices = discover_splices(fibers_a, n_fibers)
    n_real_splices = sum(1 for sp in splices if not sp['is_bend'])
    n_bend_cols    = sum(1 for sp in splices if sp['is_bend'])
    print(f'  Found {n_real_splices} real splice closures + {n_bend_cols} bend-zone columns:')
    for si, sp in enumerate(splices):
        kind = 'BEND' if sp['is_bend'] else f"S{sp['splice_num']}"
        print(f"    {kind:>6}: {sp['position_km']:8.2f} km  "
              f"({sp['position_km']*3280.84:,.0f} ft)  [{sp['count']} fibers]")

    # Span length from median EOL
    eof_vals = []
    for r in fibers_a.values():
        for e in r['events']:
            if e['is_end']:
                eof_vals.append(e['dist_km'])
                break
    span_km = round(float(np.median(eof_vals)) if eof_vals else 0, 2)
    print(f'  Span: {span_km} km ({span_km * 3280.84:,.0f} ft)')

    # ── Break detection: ribbon EOL comparison (pass 1) ──────────────────────
    print('\nDetecting breaks by ribbon EOL comparison...')
    eof_a_all  = collect_eof_positions(fibers_a)
    raw_breaks = detect_breaks_by_ribbon(eof_a_all, args.ribbon_size)

    if raw_breaks:
        print(f'  {len(raw_breaks)} broken fiber(s) detected:')
        for fnum in sorted(raw_breaks.keys()):
            ribbon_idx = (fnum - 1) // args.ribbon_size
            print(f'    Fiber {fnum} (ribbon {ribbon_idx+1}): '
                  f'EOL at {raw_breaks[fnum]:.3f} km '
                  f'({raw_breaks[fnum]*3280.84:,.0f} ft from {args.site_a})')
    else:
        print('  No breaks detected — all fibers reach the far end.')

    # ── Bend detection: segment-slope ribbon consensus (pass 2a) ────────────
    print('\nDetecting bend zones by ribbon consensus...')
    all_bends = detect_bend_zones(fibers_a, splices, args.ribbon_size,
                                  args.bend_threshold, known_breaks=raw_breaks)
    bend_fiber_count = len(all_bends)
    ribbon_seg_pairs = set()
    for fnum, segs in all_bends.items():
        ribbon_idx = (fnum - 1) // args.ribbon_size
        for s in segs:
            ribbon_seg_pairs.add((ribbon_idx, s['start_km']))
    print(f'  Segment-slope: {bend_fiber_count} fibers in '
          f'{len(ribbon_seg_pairs)} confirmed bend zone(s)')

    # ── Trace-based bend detection: sliding window on RawSamples (pass 2b) ──
    print('  Scanning RawSamples traces for localized bends...')
    splice_kms = [sp['position_km'] for sp in splices]
    trace_bends, trace_bend_cols = detect_trace_bend_zones(
        fibers_a, args.ribbon_size, args.bend_threshold,
        known_event_kms=splice_kms, known_breaks=raw_breaks,
        fibers_b=fibers_b, span_km=span_km)
    print(f'  Trace-scan: {len(trace_bends)} fibers with bend candidates, '
          f'{len(trace_bend_cols)} consensus positions')

    # Merge trace-detected bend columns into the splice list if they aren't
    # already covered by an existing column
    new_cols_added = 0
    for bc in trace_bend_cols:
        bc_km = bc['position_km']
        # Check if this position is already within 0.8 km of an existing column
        if any(abs(sp['position_km'] - bc_km) < 0.8 for sp in splices):
            continue
        splices.append(bc)
        new_cols_added += 1
    if new_cols_added:
        splices.sort(key=lambda s: s['position_km'])
        # Re-number real splices
        snum = 0
        for sp in splices:
            if not sp['is_bend']:
                snum += 1
                sp['splice_num'] = snum
        print(f'  Added {new_cols_added} new bend column(s) from trace scan')

    # Merge trace bend data into all_bends for the attenuation notes
    for fnum, blist in trace_bends.items():
        if fnum not in all_bends:
            all_bends[fnum] = []
        existing_kms = {round(b.get('start_km', b.get('km', 0)), 1)
                        for b in all_bends[fnum]}
        for b in blist:
            if round(b['km'], 1) not in existing_kms:
                all_bends[fnum].append({
                    'start_km':    b['km'] - 0.5,
                    'end_km':      b['km'] + 0.5,
                    'excess_db_km': b['excess'],
                })

    # ── Per-fiber analysis: break localization + splice losses (pass 3) ──────
    print(f'\nAnalyzing {n_fibers} fibers...')
    all_results = {}
    all_breaks  = {}
    all_missed  = {}   # fnum -> [{km, step_db}] events the OTDR missed

    for fnum in sorted(set(list(fibers_a.keys()) + list(fibers_b.keys()))):
        ra = fibers_a.get(fnum)
        rb = fibers_b.get(fnum)

        pre_break = raw_breaks.get(fnum)   # None if not broken

        splice_results, break_a, noise_km, missed_a = analyze_fiber(
            fnum, ra, rb, splices, span_km,
            args.threshold,
            precomputed_break_km=pre_break)

        for rec in splice_results:
            all_results[(fnum, rec['splice_idx'])] = rec

        if break_a:
            all_breaks[fnum] = break_a

        if missed_a:
            all_missed[fnum] = missed_a

        if fnum % 50 == 0:
            print(f'  ...fiber {fnum}/{n_fibers}')

    n_flagged   = sum(1 for r in all_results.values() if r['flagged'])
    n_bfill     = sum(1 for r in all_results.values() if r['source'] == 'B_fill' and r['primary'])
    n_bend_fibs = len(all_bends)

    # Apply ribbon consensus to missed events — suppress per-fiber noise
    all_missed = apply_ribbon_consensus_to_missed(
        all_missed, ribbon_size=args.ribbon_size, min_fibers=2)

    n_missed_total = sum(len(v) for v in all_missed.values())
    print(f'\n  Breaks detected:        {len(all_breaks)}')
    print(f'  Flagged splice events:  {n_flagged}')
    print(f'  B-fill entries used:    {n_bfill}')
    print(f'  Fibers with bend zones: {n_bend_fibs}')
    print(f'  OTDR-missed events:     {n_missed_total} across {len(all_missed)} fibers (after ribbon consensus)')

    if all_breaks:
        print('\n  Break localization summary:')
        for fnum in sorted(all_breaks.keys()):
            b   = all_breaks[fnum]
            off = f"  {b['offset_ft']:.0f}ft from nearest splice" \
                  if b['offset_ft'] is not None else ''
            sl  = f"  step={b['step_loss']:.2f}dB" if b['step_loss'] else ''
            rh  = f"  refl={b['refl_height']:.2f}dB" \
                  if b.get('refl_height') is not None else ''
            cls = f"  → {b.get('break_class', '')}"
            print(f"    Fiber {fnum}: {b['break_km']:.3f} km "
                  f"({b['break_km']*3280.84:,.0f} ft from {args.site_a})"
                  f"{off}{sl}{rh}{cls}")

    print('\nWriting Excel report...')
    write_xlsx(all_results, all_breaks, all_bends, all_missed, splices, n_fibers,
               args.output, args.site_a, args.site_b, span_km,
               args.threshold, args.ribbon_size)


if __name__ == '__main__':
    main()
